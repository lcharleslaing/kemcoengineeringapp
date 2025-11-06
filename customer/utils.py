"""Utilities for parsing KOM Excel files"""
import re
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook


def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    if isinstance(obj, (datetime,)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError(f"Type {type(obj)} not serializable")


def parse_kom_excel(file_path):
    """
    Parse KOM Excel file and return dictionary of all extracted data
    Returns JSON-serializable data (dates as ISO strings, decimals as floats)
    Includes validation warnings for missing sections
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    data = {}
    data['_validation_warnings'] = []  # Track missing sections and issues
    
    # Helper function to get cell value safely
    def get_cell(row, col):
        try:
            val = ws.cell(row=row, column=col).value
            if val is None:
                return ''
            return str(val).strip()
        except:
            return ''
    
    # Helper function to get cell value as-is (for dates, numbers)
    def get_cell_raw(row, col):
        try:
            return ws.cell(row=row, column=col).value
        except:
            return None
    
    # Helper function to find value after a label in a row
    def find_value_after_label(row, label_text, max_col=15, skip_label_patterns=True):
        """Find a label in a row and return the value in the next non-empty cell"""
        label_patterns = ['Type:', 'Qty:', 'Flow (gpm):', 'TDH (ft):', 'Emissions:', 'Pump/Grav:', "Mat'l:",
                         'Packaging:', 'Piping Material:', 'Dia(in) x L(in)', 'Valve Type:', 'Tank Mat\'l',
                         'Face Plumbing Mat\'l:', 'Gas Train Orientation:', 'PLC(s):', 'SPLIT VOLT',
                         'Size (BHP):', 'Length (ft):', 'TOTAL', 'Cap(s)?', '# Sections:', 'Diam (in):',
                         'Tubes:', 'Size:', 'Ht (ft):', 'GA:', 'Name:', 'Phone:', 'Email:', 'Company:',
                         'Address:', 'City:', 'State:', 'ZIP:', 'WHO', 'WHEN', 'Desired Delivery:', 'Consultant:',
                         'Contractor Name:', 'Freight:', 'International:', 'Shipping Instructions:', 'Electrical:',
                         'Fuel Type:', 'Gas Pressure:', 'Onsite Gas Supply Diameter', 'Sell Price:', 'Equip Cost:',
                         'Freight:', 'StartUp Cost:', 'Protect Cost', 'Net Revenue:', 'Install Cost:', '# of Trips:',
                         '# of Days:', 'CAPITAL', 'INSTALL']
        
        for col in range(1, max_col):
            cell_val = get_cell(row, col)
            if not cell_val:
                continue
            # Check if this cell contains the label we're looking for
            if label_text.lower() in cell_val.lower():
                # Make sure it's actually the label, not just containing the text
                cell_clean = cell_val.strip().rstrip(':')
                if cell_clean.lower() == label_text.lower().rstrip(':') or cell_val.lower().startswith(label_text.lower()):
                    # Look for value in next few columns (limit to 3 columns ahead)
                    for next_col in range(col + 1, min(col + 4, max_col)):
                        val = get_cell(row, next_col)
                        if not val:
                            continue
                        # Skip if it's a label or empty
                        if skip_label_patterns:
                            val_lower = val.lower().strip()
                            # Check if it matches any label pattern
                            if any(pattern.lower() in val_lower for pattern in label_patterns):
                                continue  # Skip this, it's a label
                            # Check if it ends with colon (likely a label)
                            if val_lower.endswith(':'):
                                continue
                        # Valid value found
                        if val not in ['', 'N/A', 'MISSING', 'None']:
                            return val
        return ''
    
    # Helper function to parse date - returns ISO string for JSON
    def parse_date(val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date().isoformat()
        if isinstance(val, str):
            for fmt in ('%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(val, fmt).date().isoformat()
                except ValueError:
                    pass
        return None
    
    # Helper function to parse decimal - returns float for JSON
    def parse_decimal(val):
        if val is None or val == '':
            return None
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).replace('$', '').replace(',', '').strip()
        try:
            return float(Decimal(val_str))
        except (InvalidOperation, ValueError):
            return None
    
    # Helper function to parse boolean
    def parse_bool(val):
        if not val:
            return False
        val_str = str(val).strip().upper()
        return val_str in ['YES', 'TRUE', '1', 'Y']
    
    # TO BE COMPLETED BY SALES - Row 2
    data['proposal_number'] = get_cell(2, 3)
    data['proposal_date'] = parse_date(get_cell_raw(2, 6))
    data['sales_rep'] = get_cell(2, 8)
    data['date_of_oc'] = parse_date(get_cell_raw(2, 11))
    
    # Row 4
    data['industry'] = get_cell(4, 3)
    data['industry_subcategory'] = get_cell(4, 5)
    # Discount - try multiple columns
    discount_str = None
    for col in [8, 9, 10]:
        val = get_cell(4, col)
        if val and val != 'Discount:' and '%' in val:
            discount_str = val
            break
        elif val and val.replace('.', '').replace('-', '').isdigit():
            discount_str = val
            break
    if discount_str:
        discount_str = discount_str.replace('%', '').replace(':', '').strip()
        data['discount'] = parse_decimal(discount_str)
    else:
        data['discount'] = None
    data['po_number'] = get_cell(4, 11)
    
    # Commissions - Row 6
    comm1_percent = get_cell(6, 4)
    if comm1_percent:
        comm1_percent = comm1_percent.replace('%', '').strip()
        data['comm_1_inside_percent'] = parse_decimal(comm1_percent)
    else:
        data['comm_1_inside_percent'] = None
    data['comm_1_inside_name'] = get_cell(6, 6)
    comm_outside = get_cell(6, 10)
    data['comm_outside_amount'] = parse_decimal(comm_outside)
    
    # Row 8
    comm2_percent = get_cell(8, 4)
    if comm2_percent:
        comm2_percent = comm2_percent.replace('%', '').strip()
        data['comm_2_inside_percent'] = parse_decimal(comm2_percent)
    else:
        data['comm_2_inside_percent'] = None
    data['comm_2_inside_name'] = get_cell(8, 6)
    # Comm outside name - try columns 9-12
    comm_outside_name = get_cell(8, 11) or get_cell(8, 10) or get_cell(8, 9)
    data['comm_outside_name'] = comm_outside_name if comm_outside_name and comm_outside_name != 'Name:' else ''
    
    # BILL TO - Rows 12-24 (Bill To is in cols 2-5, Ship To is in cols 6-9)
    # Row 12: Name in col 3 (after "Name:" label in col 2), Ship To name in col 7 (after "Name:" label in col 6)
    data['bill_to_name'] = get_cell(12, 3) or ''
    data['ship_to_name'] = get_cell(12, 7) or ''
    # Row 14: Phone in col 3 and 7
    data['bill_to_phone'] = get_cell(14, 3) or ''
    data['ship_to_phone'] = get_cell(14, 7) or ''
    # Row 16: Email in col 3 and 7
    data['bill_to_email'] = get_cell(16, 3) or ''
    data['ship_to_email'] = get_cell(16, 7) or ''
    # Row 18: Company in col 3 and 7
    data['bill_to_company'] = get_cell(18, 3) or ''
    data['ship_to_company'] = get_cell(18, 7) or ''
    # Row 20: Address in col 3 and 7
    data['bill_to_address'] = get_cell(20, 3) or ''
    data['ship_to_address'] = get_cell(20, 7) or ''
    # Row 22: City in col 3 and 7
    data['bill_to_city'] = get_cell(22, 3) or ''
    data['ship_to_city'] = get_cell(22, 7) or ''
    # Row 24: State in col 3, ZIP in col 4, Ship To State in col 7, ZIP in col 9
    data['bill_to_state'] = get_cell(24, 3) or ''
    data['bill_to_zip'] = get_cell(24, 4) or ''
    data['ship_to_state'] = get_cell(24, 7) or ''
    data['ship_to_zip'] = get_cell(24, 9) or ''
    
    # Tax - Row 26
    tax_exempt_str = get_cell(26, 4).upper()
    data['tax_exempt'] = (tax_exempt_str == 'YES') if tax_exempt_str else False
    exempt_cert_str = get_cell(26, 8)
    data['exempt_cert_in_hand'] = parse_bool(exempt_cert_str) if exempt_cert_str else False
    # Tax Action Who - try columns 11-14
    tax_action_who = get_cell(26, 13) or get_cell(26, 12) or get_cell(26, 11)
    data['tax_action_who'] = tax_action_who if tax_action_who and tax_action_who.upper() not in ['WHO', 'WHEN'] else ''
    # Tax Action When - try columns 12-14 in row 27
    tax_action_when = get_cell_raw(27, 13) or get_cell_raw(27, 12) or get_cell_raw(27, 11)
    data['tax_action_when'] = parse_date(tax_action_when)
    
    # Row 30
    tax_status_str = get_cell(30, 4)
    data['confirm_tax_status_noted'] = parse_bool(tax_status_str) if tax_status_str else False
    sage_str = get_cell(30, 6)
    data['customer_in_sage'] = parse_bool(sage_str) if sage_str else False
    
    # Payment Milestones - Rows 33-37
    for i in range(1, 6):
        row = 32 + i
        # Check if row 3 is just "#1:", "#2:", etc. - if so, skip it
        row_label = get_cell(row, 3)
        if row_label and row_label.strip() in ['#1:', '#2:', '#3:', '#4:', '#5:']:
            # This row is just a label, skip it
            data[f'payment_milestone_{i}_event'] = ''
            data[f'payment_milestone_{i}_percent'] = None
            data[f'payment_milestone_{i}_terms'] = ''
            data[f'payment_milestone_{i}_notes'] = ''
        else:
            milestone_event = get_cell(row, 4) or get_cell(row, 3)
            milestone_percent = get_cell_raw(row, 5) or get_cell_raw(row, 4)
            milestone_terms = get_cell(row, 6) or get_cell(row, 5)
            milestone_notes = get_cell(row, 7) or get_cell(row, 6)
            
            data[f'payment_milestone_{i}_event'] = milestone_event
            data[f'payment_milestone_{i}_percent'] = parse_decimal(milestone_percent)
            data[f'payment_milestone_{i}_terms'] = milestone_terms
            data[f'payment_milestone_{i}_notes'] = milestone_notes
    
    # Row 43 - Consultant and Contractor
    consultant = get_cell(43, 4)  # Consultant: in col 3, value in col 4
    if consultant and consultant not in ['Consultant:', 'Desired Delivery:', 'Contractor Name:', 'None'] and not consultant.endswith(':'):
        data['consultant'] = consultant
    else:
        data['consultant'] = ''
    contractor = get_cell(43, 7)  # Contractor Name: in col 6, value in col 7
    if contractor and contractor not in ['Contractor Name:', 'Desired Delivery:', 'TBD'] and not contractor.endswith(':'):
        data['contractor_name'] = contractor
    else:
        data['contractor_name'] = ''
    
    # Row 45 - Desired Delivery
    desired_delivery = get_cell(45, 4)  # Desired Delivery: in col 3, value in col 4
    if desired_delivery and 'Desired Delivery' not in desired_delivery and not desired_delivery.endswith(':'):
        data['desired_delivery'] = desired_delivery
    else:
        data['desired_delivery'] = ''
    
    # SHIPPING - Row 48
    data['freight'] = get_cell(48, 3)  # Freight: in col 2, value in col 3
    intl_str = get_cell(48, 6)  # International: in col 5, value in col 6
    if intl_str:
        intl_str = intl_str.upper()
        data['international'] = (intl_str == 'YES')
    else:
        data['international'] = False
    data['international_freight_method'] = get_cell(48, 10)  # If Int'l, Freight Method: in col 9, value in col 10
    
    # Row 50 - Shipping Instructions
    shipping_instructions = get_cell(50, 3)  # Instructions: in col 2, value in col 3
    if shipping_instructions and 'Instructions:' not in shipping_instructions:
        data['shipping_instructions'] = shipping_instructions
    else:
        data['shipping_instructions'] = ''
    
    # SPECIFICATIONS - Row 53
    specs_str = get_cell(53, 4)
    data['specifications_provided'] = parse_bool(specs_str) if specs_str else False
    agreed_str = get_cell(53, 6)
    data['specifications_agreed'] = parse_bool(agreed_str) if agreed_str else False
    ld_str = get_cell(53, 10)
    data['liquidated_damages'] = parse_bool(ld_str) if ld_str else False
    data['liquidated_damages_rate_cap'] = get_cell(54, 10)
    
    # Row 55
    passivation_str = get_cell(55, 3)
    data['passivation'] = parse_bool(passivation_str) if passivation_str else False
    data['passivation_in_out_both'] = get_cell(55, 6)
    
    # APPROVAL PRINTS - Row 57
    approval_dates_str = get_cell(57, 7)
    if approval_dates_str:
        data['approval_prints_ll_mech'] = approval_dates_str
        parts = approval_dates_str.split(';')
        elect_part = parts[1].strip() if len(parts) > 1 else ''
        if 'Elect:' in elect_part:
            data['approval_prints_elect'] = elect_part.replace('Elect:', '').strip()
        else:
            data['approval_prints_elect'] = ''
    else:
        data['approval_prints_ll_mech'] = ''
        data['approval_prints_elect'] = ''
    
    # Row 58
    approval_req_str = get_cell(58, 4)
    data['approval_prints_required'] = parse_bool(approval_req_str) if approval_req_str else False
    electrical_str = get_cell(58, 6)
    data['approval_prints_electrical'] = parse_bool(electrical_str) if electrical_str else False
    
    # Row 59
    eng_order_str = get_cell(59, 6)
    data['engineering_order_prior_to_approval'] = parse_bool(eng_order_str) if eng_order_str else False
    
    # EQUIPMENT - HTR - 1 (Find dynamically)
    # First, try to find the HTR - 1 section dynamically
    htr1_row = None
    htr1_size_row = None
    
    # Search for "HTR - 1" or "HTR-1" in the EQUIPMENT section (typically around rows 55-70)
    for row in range(55, 70):
        cell_val = get_cell(row, 1).upper()
        if 'HTR' in cell_val and ('1' in cell_val or 'ONE' in cell_val):
            htr1_row = row
            # Size row is typically 2 rows after the main row
            htr1_size_row = row + 2
            break
    
    # Fallback to default rows if not found
    if htr1_row is None:
        htr1_row = 61
        htr1_size_row = 63
        data['_validation_warnings'].append('HTR-1 section not found dynamically, using default rows')
    
    # Now extract the data - use dynamic label searching
    # Qty: can be in various columns, search for the label
    htr1_qty = None
    for col in range(1, 12):
        cell_val = get_cell(htr1_row, col)
        if cell_val and 'Qty:' in cell_val:
            # Found Qty: label, get value from next column
            qty_val = get_cell_raw(htr1_row, col + 1)
            if qty_val:
                htr1_qty = int(parse_decimal(qty_val)) if parse_decimal(qty_val) else None
            break
    # Fallback: try standard column 5
    if htr1_qty is None:
        qty_val = get_cell_raw(htr1_row, 5)
        htr1_qty = int(parse_decimal(qty_val)) if qty_val and parse_decimal(qty_val) else None
    data['htr_1_qty'] = htr1_qty
    
    # Type: search for label
    htr1_type = ''
    for col in range(1, 12):
        cell_val = get_cell(htr1_row, col)
        if cell_val and 'Type:' in cell_val:
            # Found Type: label, get value from next column
            type_val = get_cell(htr1_row, col + 1)
            if type_val and type_val not in ['Type:', 'Mat\'l:', 'Emissions:'] and not type_val.endswith(':'):
                htr1_type = type_val
            break
    # Fallback: try standard column 7
    if not htr1_type:
        type_val = get_cell(htr1_row, 7)
        htr1_type = type_val if type_val and type_val not in ['Type:', 'Mat\'l:', 'Emissions:'] and not type_val.endswith(':') else ''
    data['htr_1_type'] = htr1_type
    
    # Emissions: search for label
    htr1_emissions = ''
    for col in range(1, 12):
        cell_val = get_cell(htr1_row, col)
        if cell_val and 'Emissions:' in cell_val:
            # Found Emissions: label, get value from next column
            em_val = get_cell(htr1_row, col + 1)
            if em_val and em_val != 'Emissions:' and not em_val.endswith(':'):
                htr1_emissions = em_val
            break
    # Fallback: try standard column 10
    if not htr1_emissions:
        em_val = get_cell(htr1_row, 10)
        htr1_emissions = em_val if em_val and em_val != 'Emissions:' and not em_val.endswith(':') else ''
    data['htr_1_emissions'] = htr1_emissions
    # Size row: Search for Size:, Pump/Grav:, and Mat'l: labels dynamically
    htr1_size = None
    htr1_unit = None
    
    # Search for "Size:" label in the size row
    for col in range(1, 12):
        cell_val = get_cell(htr1_size_row, col)
        if cell_val and 'Size:' in cell_val:
            # Found Size: label, try next few columns for the value and unit
            for next_col in range(col + 1, min(col + 5, 12)):
                val = get_cell(htr1_size_row, next_col)
                if val and val not in ['Size:', 'Type:', 'Mat\'l:', 'Pump/Grav:'] and not val.endswith(':'):
                    # Check if it's a number (size value) or unit
                    if val.replace('.', '').replace('-', '').replace(' ', '').isdigit() or parse_decimal(val):
                        htr1_size = val
                        # Try next column for unit
                        unit_val = get_cell(htr1_size_row, next_col + 1)
                        if unit_val and ('x106BTU/hr' in unit_val.upper() or 'mmbtu' in unit_val.lower() or 'btu/hr' in unit_val.lower()):
                            htr1_unit = unit_val
                        break
                    elif 'x106BTU/hr' in val.upper() or 'mmbtu' in val.lower() or 'btu/hr' in val.lower():
                        # This is the unit, size might be in previous column
                        htr1_unit = val
                        prev_val = get_cell(htr1_size_row, next_col - 1)
                        if prev_val and (prev_val.replace('.', '').replace('-', '').replace(' ', '').isdigit() or parse_decimal(prev_val)):
                            htr1_size = prev_val
                        break
            break
    
    # Fallback: try standard locations
    if not htr1_size:
        htr1_size = get_cell(htr1_size_row, 3)
        htr1_unit = get_cell(htr1_size_row, 4)
        if htr1_size and 'Size:' in htr1_size:
            htr1_size = get_cell(htr1_size_row, 4)
            htr1_unit = get_cell(htr1_size_row, 5)
    
    # Compose the size string
    if htr1_size and htr1_size not in ['Size:', 'Type:', 'Mat\'l:'] and not htr1_size.endswith(':'):
        # Clean up the size value - handle cases like "1.2 x10⁶" or "1.2 x106"
        size_clean = htr1_size.strip()
        if parse_decimal(size_clean) or size_clean.replace('.', '').replace('-', '').replace(' ', '').isdigit():
            composed = f"{size_clean} {htr1_unit}".strip() if htr1_unit else size_clean.strip()
            # Avoid saving unit-only string
            if composed.strip().lower() in ['x106btu/hr', 'x10⁶btu/hr', 'mmbtu']:
                data['htr_1_size'] = ''
            else:
                data['htr_1_size'] = composed
        else:
            data['htr_1_size'] = ''
    else:
        data['htr_1_size'] = ''
    
    # Pump/Grav: search for label
    htr1_pump = ''
    for col in range(1, 12):
        cell_val = get_cell(htr1_size_row, col)
        if cell_val and ('Pump/Grav:' in cell_val or 'Pump/Grav' in cell_val):
            # Found Pump/Grav: label, get value from next column
            pump_val = get_cell(htr1_size_row, col + 1)
            if pump_val and pump_val not in ['Pump/Grav:', 'Type:', 'Mat\'l:'] and not pump_val.endswith(':'):
                htr1_pump = pump_val
            break
    # Fallback: try standard column 6
    if not htr1_pump:
        pump_val = get_cell(htr1_size_row, 6)
        htr1_pump = pump_val if pump_val and pump_val not in ['Pump/Grav:', 'Type:', 'Mat\'l:'] and not pump_val.endswith(':') else ''
    data['htr_1_pump_grav'] = htr1_pump
    
    # Mat'l: search for label
    htr1_mat = ''
    for col in range(1, 12):
        cell_val = get_cell(htr1_size_row, col)
        if cell_val and ("Mat'l:" in cell_val or "Mat'l" in cell_val or "Material:" in cell_val):
            # Found Mat'l: label, get value from next column
            mat_val = get_cell(htr1_size_row, col + 1)
            if mat_val and mat_val not in ["Mat'l:", 'Type:', 'Emissions:'] and not mat_val.endswith(':'):
                htr1_mat = mat_val
            break
    # Fallback: try standard column 8
    if not htr1_mat:
        mat_val = get_cell(htr1_size_row, 8)
        htr1_mat = mat_val if mat_val and mat_val not in ["Mat'l:", 'Type:', 'Emissions:'] and not mat_val.endswith(':') else ''
    data['htr_1_material'] = htr1_mat
    
    # HTR - 2 (Row 65 is main row, row 67 is size row - but only if HTR - 2 exists)
    # First check if HTR - 2 section exists
    htr2_exists = False
    for row in range(64, 68):
        if 'HTR - 2' in get_cell(row, 1).upper():
            htr2_exists = True
            break
    
    if htr2_exists:
        # HTR - 2 exists: Row 65 is main row, Row 67 is size row
        htr2_row = 65
        htr2_size_row = 67
        htr2_qty = get_cell_raw(htr2_row, 5)
        data['htr_2_qty'] = int(parse_decimal(htr2_qty)) if htr2_qty and parse_decimal(htr2_qty) else None
        htr2_type = get_cell(htr2_row, 7)
        data['htr_2_type'] = htr2_type if htr2_type and htr2_type not in ['Type:', 'Mat\'l:', 'Emissions:'] and not htr2_type.endswith(':') else ''
        htr2_emissions = get_cell(htr2_row, 10)
        data['htr_2_emissions'] = htr2_emissions if htr2_emissions and htr2_emissions != 'Emissions:' and not htr2_emissions.endswith(':') else ''
        htr2_size_val = get_cell(htr2_size_row, 3)
        htr2_unit_val = get_cell(htr2_size_row, 4)
        if htr2_size_val and htr2_size_val not in ['Size:', 'Size (BHP):', 'Pump/Grav:', 'Mat\'l:'] and not htr2_size_val.endswith(':'):
            composed2 = f"{htr2_size_val} {htr2_unit_val}".strip()
            data['htr_2_size'] = '' if composed2.strip().lower() == 'x106btu/hr' else composed2
        else:
            data['htr_2_size'] = ''
        htr2_pump = get_cell(htr2_size_row, 6)
        data['htr_2_pump_grav'] = htr2_pump if htr2_pump and htr2_pump not in ['Pump/Grav:', "Mat'l:", 'Size:', 'Size (BHP):'] and not htr2_pump.endswith(':') else ''
        htr2_mat = get_cell(htr2_size_row, 8)
        data['htr_2_material'] = htr2_mat if htr2_mat and htr2_mat not in ["Mat'l:", 'Type:', 'Emissions:', 'Pump/Grav:'] and not htr2_mat.endswith(':') else ''
    else:
        # HTR - 2 does not exist, set all fields to empty/None
        data['htr_2_qty'] = None
        data['htr_2_type'] = ''
        data['htr_2_emissions'] = ''
        data['htr_2_size'] = ''
        data['htr_2_pump_grav'] = ''
        data['htr_2_material'] = ''
    
    # STK ECON (Row 69)
    stk_size = get_cell(69, 4) or get_cell(69, 3)
    data['stk_econ_size_bhp'] = stk_size if stk_size and stk_size not in ['Size (BHP):', 'TOTAL', 'Cap(s)?', 'Tubes:'] and not stk_size.endswith(':') else ''
    stk_pump = get_cell(69, 7) or get_cell(69, 6)
    data['stk_econ_pump_grav'] = stk_pump if stk_pump and stk_pump not in ['Pump/Grav:', 'TOTAL', 'Cap(s)?', 'Tubes:'] and not stk_pump.endswith(':') else ''
    stk_mat = get_cell(69, 9) or get_cell(69, 8)
    data['stk_econ_material'] = stk_mat if stk_mat and stk_mat not in ["Mat'l:", 'TOTAL', 'Cap(s)?', 'Tubes:'] and not stk_mat.endswith(':') else ''
    
    # STACK(S) (Row 71)
    data['stack_length_ft'] = parse_decimal(get_cell(71, 5))  # Length (ft): in col 4, value in col 5
    stack_total = get_cell(71, 7)  # TOTAL is in col 6, value in col 7
    if stack_total and stack_total.upper() not in ['TOTAL', 'YES', 'NO'] and stack_total not in ['Tubes:', 'TOTAL', 'Cap(s)?'] and not stack_total.endswith(':'):
        data['stack_total'] = stack_total
    else:
        data['stack_total'] = ''
    caps_str = get_cell(71, 10)  # Cap(s)? in col 9, value in col 10
    data['stack_caps'] = parse_bool(caps_str) if caps_str and caps_str not in ['TOTAL', 'Tubes:', 'Cap(s)?'] else False
    
    # HR (Row 73)
    hr_sections = get_cell(73, 5)  # # Sections: in col 4, value in col 5
    data['hr_sections'] = int(parse_decimal(hr_sections)) if hr_sections and parse_decimal(hr_sections) else None
    hr_diam = get_cell(73, 7)  # Diam (in): in col 6, value in col 7
    data['hr_diam_in'] = parse_decimal(hr_diam) if hr_diam and 'Diam (in):' not in hr_diam else None
    hr_tubes = get_cell(73, 9)  # Tubes: in col 8, value in col 9
    data['hr_tubes'] = hr_tubes if hr_tubes and 'Tubes:' not in hr_tubes else ''
    hr_mat = get_cell(73, 11)  # Mat'l: in col 10, value in col 11
    data['hr_material'] = hr_mat if hr_mat and "Mat'l:" not in hr_mat else ''
    
    # TANKS (Rows 77, 79, 81 - each tank is on its own row)
    # Row 77: #1:,Type:,HW,Dia (in):,84,Ht (ft):,15,GA:,STD,Mat'l:,304
    # So Type value is col 4, Dia value is col 6, Ht value is col 8, GA value is col 10, Mat'l value is col 12
    # First, try to find the TANKS section dynamically
    tanks_section_row = None
    for row in range(70, 85):
        if 'TANKS' in get_cell(row, 1).upper():
            tanks_section_row = row
            break
    
    if not tanks_section_row:
        tanks_section_row = 75  # Default fallback
        data['_validation_warnings'].append('TANKS section not found, using default row 75')
    
    # Tanks are typically 2 rows after the TANKS header, then every 2 rows
    # Based on report: tanks consistently at rows 77, 79, 81 with markers #1:, #2:, #3: in column 1
    tanks_found = 0
    for i in range(1, 4):
        # Try both fixed row numbers and dynamic positioning
        tank_row = tanks_section_row + 2 + ((i - 1) * 2)  # 77, 79, 81 if section is at 75
        # Also try the fixed row numbers as fallback
        tank_row_fallback = 76 + (i * 2)  # 77, 79, 81
        
        # Check if this row has a tank number marker (#1:, #2:, #3:)
        row_marker = get_cell(tank_row, 1)
        if not row_marker or not row_marker.startswith('#'):
            # Try fallback row
            row_marker = get_cell(tank_row_fallback, 1)
            if row_marker and row_marker.startswith('#'):
                tank_row = tank_row_fallback
        
        # Only process if we found a tank marker
        if row_marker and row_marker.startswith('#'):
            tanks_found += 1
            # Some files place values at cols 4/6/8/10/12, others at 3/5/7/9/11
            # Try primary set first, then fallback
            def pick_text(col_primary, col_fallback):
                val = get_cell(tank_row, col_primary)
                if val and val not in ['Type:', 'Dia (in):', 'Ht (ft):', 'GA:', "Mat'l:"] and not val.endswith(':'):
                    return val
                val_fb = get_cell(tank_row, col_fallback)
                return val_fb if val_fb and val_fb not in ['Type:', 'Dia (in):', 'Ht (ft):', 'GA:', "Mat'l:"] and not val_fb.endswith(':') else ''

            def pick_number(col_primary, col_fallback):
                # Try primary column first
                raw = get_cell_raw(tank_row, col_primary)
                num = parse_decimal(raw) if raw is not None else None
                
                # If primary didn't work, try fallback
                if num is None:
                    raw_fb = get_cell_raw(tank_row, col_fallback)
                    num = parse_decimal(raw_fb) if raw_fb is not None else None
                
                # Also try adjacent columns in case the structure is slightly different
                if num is None:
                    for adj_col in [col_primary - 1, col_primary + 1, col_fallback - 1, col_fallback + 1]:
                        if adj_col > 0:
                            raw_adj = get_cell_raw(tank_row, adj_col)
                            num_adj = parse_decimal(raw_adj) if raw_adj is not None else None
                            if num_adj is not None:
                                num = num_adj
                                break
                
                # Return the number, preserving decimals if needed, or None if not found
                if num is not None:
                    # For diameter and height, we want integers, but allow decimals
                    if isinstance(num, (int, float)):
                        # If it's a whole number, return as int, otherwise keep as float
                        return int(num) if num == int(num) else float(num)
                return None

            tank_type_val = pick_text(4, 3)
            data[f'tank_{i}_type'] = tank_type_val

            # For diameter, try multiple approaches:
            # 1. Standard column positions
            dia_val = pick_number(6, 5)
            # 2. If that didn't work, search for "Dia (in):" label dynamically
            if dia_val is None:
                dia_label_val = find_value_after_label(tank_row, 'Dia (in):', max_col=15)
                if dia_label_val:
                    dia_val = parse_decimal(dia_label_val)
            # 3. Also check if the value is in the cell right after "Dia (in):" label
            if dia_val is None:
                for col in range(1, 15):
                    cell_val = get_cell(tank_row, col)
                    if 'Dia (in):' in cell_val or 'Dia(in):' in cell_val:
                        # Check next few columns for the value
                        for next_col in range(col + 1, min(col + 4, 15)):
                            next_val = get_cell_raw(tank_row, next_col)
                            if next_val is not None:
                                parsed = parse_decimal(next_val)
                                if parsed is not None:
                                    dia_val = parsed
                                    break
                        if dia_val is not None:
                            break
            
            data[f'tank_{i}_dia_in'] = int(dia_val) if dia_val is not None and dia_val == int(dia_val) else (float(dia_val) if dia_val is not None else None)
            
            # For height, same approach
            ht_val = pick_number(8, 7)
            if ht_val is None:
                ht_label_val = find_value_after_label(tank_row, 'Ht (ft):', max_col=15)
                if ht_label_val:
                    ht_val = parse_decimal(ht_label_val)
            if ht_val is None:
                for col in range(1, 15):
                    cell_val = get_cell(tank_row, col)
                    if 'Ht (ft):' in cell_val or 'Ht(ft):' in cell_val:
                        for next_col in range(col + 1, min(col + 4, 15)):
                            next_val = get_cell_raw(tank_row, next_col)
                            if next_val is not None:
                                parsed = parse_decimal(next_val)
                                if parsed is not None:
                                    ht_val = parsed
                                    break
                        if ht_val is not None:
                            break
            
            data[f'tank_{i}_ht_ft'] = int(ht_val) if ht_val is not None and ht_val == int(ht_val) else (float(ht_val) if ht_val is not None else None)

            tank_ga_val = pick_text(10, 9)
            data[f'tank_{i}_ga'] = tank_ga_val

            tank_mat_val = pick_text(12, 11)
            data[f'tank_{i}_material'] = tank_mat_val
        else:
            # No tank found, set to empty
            data[f'tank_{i}_type'] = ''
            data[f'tank_{i}_dia_in'] = None
            data[f'tank_{i}_ht_ft'] = None
            data[f'tank_{i}_ga'] = ''
            data[f'tank_{i}_material'] = ''
    
    if tanks_found == 0:
        data['_validation_warnings'].append('No tanks found with markers #1:, #2:, #3:')
    
    # PUMPS - Find dynamically like tanks
    pumps_section_row = None
    for row in range(80, 95):
        if 'PUMPS' in get_cell(row, 1).upper():
            pumps_section_row = row
            break
    
    if not pumps_section_row:
        pumps_section_row = 82  # Default fallback
        data['_validation_warnings'].append('PUMPS section not found, using default row 82')
    
    # Packaging and Piping Material are typically 1-2 rows after PUMPS header
    pump_packaging_row = pumps_section_row + 1
    pump_pack = find_value_after_label(pump_packaging_row, 'Packaging:', max_col=15)
    if not pump_pack:
        # Try direct column access
        pump_pack = get_cell(pump_packaging_row, 6)  # Packaging: in col 5, value in col 6
        if pump_pack in ['Packaging:', 'Piping Material:', 'Type:', 'Qty:'] or pump_pack.endswith(':'):
            pump_pack = ''
    data['pump_packaging'] = pump_pack if pump_pack else ''
    
    pump_pipe = find_value_after_label(pump_packaging_row, 'Piping Material:', max_col=15)
    if not pump_pipe:
        # Try direct column access
        pump_pipe = get_cell(pump_packaging_row, 8)  # Piping Material: in col 7, value in col 8
        if pump_pipe in ['Piping Material:', 'Packaging:', 'Type:', 'Qty:'] or pump_pipe.endswith(':'):
            pump_pipe = ''
    data['pump_piping_material'] = pump_pipe if pump_pipe else ''
    
    # Find pump rows by looking for #1:, #2:, #3:, #4: markers
    pumps_found = 0
    for i in range(1, 5):
        # Try both fixed row numbers and dynamic positioning
        pump_row = pumps_section_row + 2 + ((i - 1) * 2)  # Typically 2 rows after PUMPS, then every 2 rows
        pump_row_fallback = 84 + (i * 2)  # 85, 87, 89, 91
        
        # Check if this row has a pump number marker (#1:, #2:, #3:, #4:)
        row_marker = get_cell(pump_row, 1)
        if not row_marker or not (row_marker.startswith('#') and str(i) in row_marker):
            # Try fallback row
            row_marker = get_cell(pump_row_fallback, 1)
            if row_marker and row_marker.startswith('#') and str(i) in row_marker:
                pump_row = pump_row_fallback
        
        # Only process if we found a pump marker
        if row_marker and row_marker.startswith('#') and str(i) in row_marker:
            pumps_found += 1
            
            # Type: in col 4, value in col 5; Qty: in col 6, value in col 7; Flow (gpm): in col 8, value in col 9; TDH (ft): in col 10, value in col 11
            # Try multiple approaches for each field
            pump_type = find_value_after_label(pump_row, 'Type:', max_col=15)
            if not pump_type:
                pump_type = get_cell(pump_row, 5)
                if pump_type == 'Type:' or pump_type.endswith(':'):
                    pump_type = ''
            data[f'pump_{i}_type'] = pump_type if pump_type else ''
            
            # Qty
            pump_qty_val = find_value_after_label(pump_row, 'Qty:', max_col=15)
            if pump_qty_val:
                pump_qty = parse_decimal(pump_qty_val)
            else:
                pump_qty_raw = get_cell_raw(pump_row, 7)
                pump_qty = parse_decimal(pump_qty_raw) if pump_qty_raw is not None else None
            data[f'pump_{i}_qty'] = int(pump_qty) if pump_qty is not None and pump_qty == int(pump_qty) else (float(pump_qty) if pump_qty is not None else None)
            
            # Flow (gpm)
            pump_flow_val = find_value_after_label(pump_row, 'Flow (gpm):', max_col=15)
            if pump_flow_val:
                pump_flow = parse_decimal(pump_flow_val)
            else:
                pump_flow_raw = get_cell_raw(pump_row, 9)
                pump_flow = parse_decimal(pump_flow_raw) if pump_flow_raw is not None else None
            data[f'pump_{i}_flow_gpm'] = int(pump_flow) if pump_flow is not None and pump_flow == int(pump_flow) else (float(pump_flow) if pump_flow is not None else None)
            
            # TDH (ft)
            pump_tdh_val = find_value_after_label(pump_row, 'TDH (ft):', max_col=15)
            if pump_tdh_val:
                pump_tdh = parse_decimal(pump_tdh_val)
            else:
                pump_tdh_raw = get_cell_raw(pump_row, 11)
                pump_tdh = parse_decimal(pump_tdh_raw) if pump_tdh_raw is not None else None
            data[f'pump_{i}_tdh_ft'] = int(pump_tdh) if pump_tdh is not None and pump_tdh == int(pump_tdh) else (float(pump_tdh) if pump_tdh is not None else None)
        else:
            # No pump found, set to empty
            data[f'pump_{i}_type'] = ''
            data[f'pump_{i}_qty'] = None
            data[f'pump_{i}_flow_gpm'] = None
            data[f'pump_{i}_tdh_ft'] = None
    
    if pumps_found == 0:
        data['_validation_warnings'].append('No pumps found with markers #1:, #2:, #3:, #4:')
    
    # Steam Heaters (Rows 94-96)
    for i in range(1, 3):
        row = 92 + (i * 2)  # 94, 96
        dia_val = get_cell(row, 5) or get_cell(row, 4)
        if dia_val and 'x' in dia_val:
            parts = dia_val.split('x')
            if len(parts) >= 2:
                dia_num = re.sub(r'[^\d.]', '', parts[0].strip())
                length_num = re.sub(r'[^\d.]', '', parts[1].strip())
                data[f'steam_heater_{i}_dia_in'] = parse_decimal(dia_num)
                data[f'steam_heater_{i}_length_in'] = parse_decimal(length_num)
            else:
                data[f'steam_heater_{i}_dia_in'] = None
                data[f'steam_heater_{i}_length_in'] = None
        else:
            data[f'steam_heater_{i}_dia_in'] = None
            data[f'steam_heater_{i}_length_in'] = None
        steam_mat = get_cell(row, 7) or get_cell(row, 6)
        data[f'steam_heater_{i}_material'] = steam_mat if steam_mat and "Mat'l:" not in steam_mat and not steam_mat.endswith(':') else ''
        steam_valve = get_cell(row, 10) or get_cell(row, 9)
        data[f'steam_heater_{i}_valve_type'] = steam_valve if steam_valve and 'Valve Type:' not in steam_valve and not steam_valve.endswith(':') else ''
    
    # SOFTENER (Row 99)
    asme_str = get_cell(99, 5) or get_cell(99, 4) or get_cell(99, 3)
    data['softener_asme_coded'] = parse_bool(asme_str) if asme_str else False
    softener_tank = get_cell(99, 8) or get_cell(99, 7) or get_cell(99, 6)
    data['softener_tank_material'] = softener_tank if softener_tank and 'Tank Mat\'l' not in softener_tank and not softener_tank.endswith(':') else ''
    softener_face = get_cell(99, 11) or get_cell(99, 10) or get_cell(99, 9)
    data['softener_face_plumbing_material'] = softener_face if softener_face and 'Face Plumbing Mat\'l:' not in softener_face and not softener_face.endswith(':') else ''
    
    # PANEL(S) (Row 100/102)
    panel_qty = get_cell_raw(100, 4) or get_cell_raw(100, 3) or get_cell_raw(102, 4) or get_cell_raw(102, 3)
    data['panel_qty'] = int(parse_decimal(panel_qty)) if panel_qty and parse_decimal(panel_qty) else None
    panel_plc = get_cell(100, 6) or get_cell(100, 5) or get_cell(102, 7) or get_cell(102, 6)
    data['panel_plc'] = panel_plc if panel_plc and 'PLC(s):' not in panel_plc and not panel_plc.endswith(':') else ''
    panel_split = get_cell(100, 8) or get_cell(100, 7) or get_cell(102, 11) or get_cell(102, 10) or get_cell(102, 9)
    data['panel_split_volt'] = panel_split if panel_split and panel_split != 'SPLIT VOLT' and not panel_split.endswith(':') else ''
    
    # OTHER (Row 105)
    vent_str = get_cell(105, 2)
    data['other_vent_condenser'] = 'Vent Condenser' in vent_str or bool(vent_str)
    shaker_str = get_cell(105, 4)
    data['other_shaker_screen'] = 'Shaker Screen' in shaker_str or bool(shaker_str)
    
    # UTILITIES (Row 107: City Water Meter, Row 109: Electrical/Fuel/Gas Pressure, Row 111: Gas Supply/Orientation, Row 113: Match Proposal)
    # Row 109: ,Electrical:,480/3/60,,Fuel Type:,NG,,Gas Pressure: ,2-5,(psi)
    city_water = get_cell(107, 9)  # City Water Meter (in.): in col 8, value in col 9
    data['city_water_meter_in'] = parse_decimal(city_water) if city_water and 'City Water Meter' not in city_water else None
    electrical = get_cell(109, 3)  # Electrical: in col 2, value in col 3
    data['electrical'] = electrical if electrical and 'Electrical:' not in electrical and not electrical.endswith(':') else ''
    fuel_type = get_cell(109, 6)  # Fuel Type: in col 5, value in col 6
    data['fuel_type'] = fuel_type if fuel_type and 'Fuel Type:' not in fuel_type and not fuel_type.endswith(':') else ''
    gas_pressure = get_cell(109, 8)  # Gas Pressure: in col 7, value in col 8
    gas_pressure_unit = get_cell(109, 9)  # Unit (psi) in col 9
    if gas_pressure and gas_pressure not in ['Gas Pressure:', 'Gas Train Orientation:', 'L/R', 'RIGHT', 'LEFT'] and not gas_pressure.endswith(':'):
        if gas_pressure_unit and gas_pressure_unit != gas_pressure and gas_pressure_unit not in ['Gas Pressure:', 'Gas Train Orientation:', '(psi)']:
            data['gas_pressure_psi'] = f"{gas_pressure} {gas_pressure_unit}".strip()
        else:
            data['gas_pressure_psi'] = gas_pressure
    else:
        data['gas_pressure_psi'] = ''
    gas_dia = get_cell(111, 9)  # Onsite Gas Supply Diameter (in.): in col 8, value in col 9
    data['onsite_gas_supply_diameter_in'] = parse_decimal(gas_dia) if gas_dia and 'Onsite Gas Supply Diameter' not in gas_dia else None
    gas_orient = get_cell(111, 10)  # Gas Train Orientation: in col 9, value in col 10
    data['gas_train_orientation'] = gas_orient if gas_orient and 'Gas Train Orientation:' not in gas_orient and not gas_orient.endswith(':') else ''
    match_str = get_cell(113, 8)  # Does this match... in col 7, value in col 8
    data['utilities_match_proposal'] = parse_bool(match_str) if match_str else False
    
    # OTHER INFO (Row 114/116-118)
    notes = get_cell(114, 3) or get_cell(114, 2) or get_cell(116, 3) or get_cell(116, 2)
    other_info = get_cell(118, 2)
    data['notes'] = notes if notes and 'Notes:' not in notes else ''
    data['other_info'] = other_info if other_info else ''
    
    # Project Info (Row 123)
    data['project_name'] = get_cell(123, 4)  # Project Name: in col 3, value in col 4
    data['project_type'] = get_cell(123, 8)  # Project Type: in col 7, value in col 8
    
    # TO BE COMPLETED BY APPS - Line Items (Row 126 is header, Row 127+ are items)
    # Based on report: start rows vary (122, 123, 125, 127), end rows vary (133, 134, 137, 138, 141)
    line_items = []
    # Find the start of line items section
    line_items_start = None
    for row in range(120, 135):
        if 'TO BE COMPLETED BY APPS' in get_cell(row, 1).upper():
            line_items_start = row + 2  # Skip header row and empty row
            break
    
    if not line_items_start:
        line_items_start = 127  # Default fallback
        data['_validation_warnings'].append('TO BE COMPLETED BY APPS section not found, using default row 127')
    
    # Find where line items end (LABOR HOURS section)
    # Based on report: LABOR HOURS found at rows 133, 134, 137, 138, 141
    line_items_end = None
    for row in range(line_items_start, 145):
        if 'LABOR HOURS' in get_cell(row, 1).upper() or 'LABOR HOURS' in get_cell(row, 2).upper():
            line_items_end = row
            break
    
    if not line_items_end:
        line_items_end = 141  # Default fallback
        data['_validation_warnings'].append('LABOR HOURS section not found, using default row 141')
    
    for row in range(line_items_start, line_items_end):
        item_num = get_cell(row, 1)
        # Row 127: 35371-01,,,"18 MMBTU DCWH, TE+, Pumped",,49550
        # Row 131: 35371-05,,,Hot Water Makeup Valve Nest,,5490.5,,,5075,415.5
        # Description is in col 4
        # Values: col 6 has value 1, but looking at row 131, value 3 is in col 9, value 4 is in col 10
        # Actually, let me check: row 131 has 5490.5 in col 6, then 5075 in col 9, 415.5 in col 10
        # So values might be: val1=col 6, val2=col 7, val3=col 9, val4=col 10 (skipping col 8?)
        # Or maybe: val1=col 6, val2=empty, val3=col 9, val4=col 10
        description = get_cell(row, 4)
        val1 = get_cell_raw(row, 6)  # Value 1 is in col 6
        val2 = get_cell_raw(row, 7)  # Value 2 is in col 7 (often empty)
        val3 = get_cell_raw(row, 9)  # Value 3 is in col 9 (not 8!)
        val4 = get_cell_raw(row, 10)  # Value 4 is in col 10 (not 9!)
        
        # Skip section headers and empty rows
        if item_num and item_num.upper() in ['LABOR HOURS', 'ITEM', 'DESCRIPTION', 'CAPITAL', 'INSTALL', 'TO BE COMPLETED BY APPS']:
            continue
        if item_num and (item_num.endswith(':') or item_num in ['Sell Price:', 'Equip Cost:', 'Freight:', 'StartUp Cost:', 'Protect Cost', 'Net Revenue:', 'Install Cost:', '# of Trips:', '# of Days:']):
            continue
        if not item_num or not item_num.strip():
            continue
            
        if item_num or description or val1 or val2 or val3 or val4:
            line_items.append({
                'item_number': item_num,
                'description': description if description and not description.isdigit() and not description.endswith(':') else '',
                'value_1': parse_decimal(val1),
                'value_2': parse_decimal(val2),
                'value_3': parse_decimal(val3),
                'value_4': parse_decimal(val4),
            })
    
    data['line_items'] = line_items
    
    # LABOR HOURS (Row 142: HR, 143: Pkg, 144: Fab, 145: Wiring)
    # Row 142: ,HR:,, -   ,,,,QTY,KN,DESCRIPTION
    # Row 143: ,Pkg:,, 213.0 ,,,Burner:,2,890-01-012,"BURNER, GAS, EB-7"
    # Values are in col 4 (after the label in col 3)
    # Find LABOR HOURS section dynamically (varies: 133, 134, 137, 138, 141)
    labor_hours_row = None
    for row in range(130, 145):
        if 'LABOR HOURS' in get_cell(row, 1).upper() or 'LABOR HOURS' in get_cell(row, 2).upper():
            labor_hours_row = row
            break
    
    if not labor_hours_row:
        labor_hours_row = 141  # Default fallback
        data['_validation_warnings'].append('LABOR HOURS section not found for labor hours extraction, using default row 141')
    
    # HR is typically on the row after LABOR HOURS, Pkg/Fab/Wiring follow
    hr_row = labor_hours_row + 1
    pkg_row = labor_hours_row + 2
    fab_row = labor_hours_row + 3
    wiring_row = labor_hours_row + 4
    
    hr_labor = get_cell(hr_row, 4) or get_cell(142, 4)
    data['labor_hr'] = hr_labor if hr_labor and hr_labor != 'HR:' and not hr_labor.endswith(':') else '-'
    pkg_val = get_cell_raw(pkg_row, 4) or get_cell_raw(143, 4)
    data['labor_pkg'] = parse_decimal(pkg_val) if pkg_val and str(pkg_val) != 'Pkg:' else None
    fab_val = get_cell_raw(fab_row, 4) or get_cell_raw(144, 4)
    data['labor_fab'] = parse_decimal(fab_val) if fab_val and str(fab_val) != 'Fab:' else None
    wiring_val = get_cell_raw(wiring_row, 4) or get_cell_raw(145, 4)
    data['labor_wiring'] = parse_decimal(wiring_val) if wiring_val and str(wiring_val) != 'Wiring:' else None
    
    # EQUIPMENT REQUIRED (Rows 143-145: Pkg, Fab, Wiring rows have equipment)
    # Based on user feedback: columns are shifted - Qty is empty, KN has Qty, Description has KN
    # So we need to read: Qty from col 8 (where KN was), KN from col 9 (where Description was), Description from col 10
    # Use dynamic rows based on LABOR HOURS position
    equipment_required = []
    
    # Helper function to extract equipment data with column shift handling
    def extract_equipment(row, equipment_type):
        # Try multiple column positions to handle variations
        # Standard: Qty in col 7, KN in col 8, Desc in col 9
        # Shifted: Qty in col 8, KN in col 9, Desc in col 10
        # Try standard first
        qty_std = get_cell_raw(row, 7)
        kn_std = get_cell(row, 8)
        desc_std = get_cell(row, 9)
        
        # Try shifted columns
        qty_shift = get_cell_raw(row, 8)
        kn_shift = get_cell(row, 9)
        desc_shift = get_cell(row, 10)
        
        # Determine which set to use - if col 7 has a number and looks like qty, use standard
        # If col 7 is empty/not a number but col 8 has a number, use shifted
        qty_val = None
        kn_val = ''
        desc_val = ''
        
        # Check if standard columns have valid data
        qty_std_parsed = parse_decimal(qty_std) if qty_std is not None else None
        if qty_std_parsed is not None and qty_std_parsed > 0:
            # Standard columns seem valid
            qty_val = int(qty_std_parsed) if qty_std_parsed == int(qty_std_parsed) else float(qty_std_parsed)
            kn_val = kn_std if kn_std and kn_std not in ['KN', 'QTY', 'DESCRIPTION'] else ''
            desc_val = desc_std if desc_std and desc_std not in ['DESCRIPTION'] else ''
        else:
            # Try shifted columns
            qty_shift_parsed = parse_decimal(qty_shift) if qty_shift is not None else None
            if qty_shift_parsed is not None and qty_shift_parsed > 0:
                # Shifted columns have the data
                qty_val = int(qty_shift_parsed) if qty_shift_parsed == int(qty_shift_parsed) else float(qty_shift_parsed)
                kn_val = kn_shift if kn_shift and kn_shift not in ['KN', 'QTY', 'DESCRIPTION'] else ''
                desc_val = desc_shift if desc_shift and desc_shift not in ['DESCRIPTION'] else ''
            else:
                # Try to find by label search
                qty_label = find_value_after_label(row, 'Qty:', max_col=15)
                if qty_label:
                    qty_val = int(parse_decimal(qty_label)) if parse_decimal(qty_label) else None
                kn_label = find_value_after_label(row, 'KN:', max_col=15)
                if kn_label:
                    kn_val = kn_label
                desc_label = find_value_after_label(row, 'Description:', max_col=15)
                if desc_label:
                    desc_val = desc_label
        
        if qty_val or kn_val or desc_val:
            return {
                'equipment_type': equipment_type,
                'qty': qty_val,
                'kn_number': kn_val,
                'description': desc_val,
            }
        return None
    
    # Burner (same row as Pkg)
    burner_row = pkg_row
    burner_data = extract_equipment(burner_row, 'Burner')
    if burner_data:
        equipment_required.append(burner_data)
    
    # Blower (same row as Fab)
    blower_row = fab_row
    blower_data = extract_equipment(blower_row, 'Blower')
    if blower_data:
        equipment_required.append(blower_data)
    
    # Media (same row as Wiring)
    media_row = wiring_row
    media_data = extract_equipment(media_row, 'Media')
    if media_data:
        equipment_required.append(media_data)
    
    data['equipment_required'] = equipment_required
    
    # CAPITAL (Row 149: Sell Price, 151: Equip Cost, 153: Freight, 155: StartUp Cost, 156: Protect Cost, 157: Net Revenue)
    # Based on report: CAPITAL section varies (rows 139, 140, 143, 144, 147)
    # Find CAPITAL section dynamically
    capital_section_row = None
    for row in range(135, 150):
        if 'CAPITAL' in get_cell(row, 1).upper():
            capital_section_row = row
            break
    
    if not capital_section_row:
        capital_section_row = 147  # Default fallback
        data['_validation_warnings'].append('CAPITAL section not found, using default row 147')
    
    # Capital fields are typically 2, 4, 6, 8 rows after CAPITAL header
    capital_sell = get_cell(capital_section_row + 2, 3) or get_cell(149, 3)  # Sell Price: in col 2, value in col 3
    data['capital_sell_price'] = parse_decimal(capital_sell) if capital_sell and 'Sell Price:' not in capital_sell else None
    capital_equip = get_cell(capital_section_row + 4, 3) or get_cell(151, 3)  # Equip Cost: in col 2, value in col 3
    data['capital_equip_cost'] = parse_decimal(capital_equip) if capital_equip and 'Equip Cost:' not in capital_equip else None
    capital_freight = get_cell(capital_section_row + 6, 3) or get_cell(153, 3)  # Freight: in col 2, value in col 3
    data['capital_freight'] = parse_decimal(capital_freight) if capital_freight and 'Freight:' not in capital_freight else None
    startup = get_cell(capital_section_row + 8, 4) or get_cell(155, 4)  # StartUp Cost: in col 3, value in col 4
    data['capital_startup_cost'] = parse_decimal(startup) if startup and 'StartUp Cost:' not in startup else None
    protect = get_cell(capital_section_row + 9, 3) or get_cell(156, 3)  # Protect Cost: in col 2, value in col 3 (or might be empty)
    data['capital_protect_cost'] = parse_decimal(protect) if protect and 'Protect Cost' not in protect else None
    net_rev = get_cell(capital_section_row + 10, 3) or get_cell(157, 3)  # Net Revenue: in col 2, value in col 3
    data['capital_net_revenue'] = parse_decimal(net_rev) if net_rev and 'Net Revenue:' not in net_rev else None
    
    # INSTALL (Rows 138-146/149-157, different columns)
    install_sell = get_cell(138, 9) or get_cell(138, 8) or get_cell(149, 10) or get_cell(149, 9) or get_cell(149, 8)
    data['install_sell_price'] = parse_decimal(install_sell) if install_sell and 'Sell Price:' not in install_sell else None
    install_cost = get_cell(140, 9) or get_cell(140, 8) or get_cell(151, 10) or get_cell(151, 9) or get_cell(151, 8)
    if install_cost and 'Install Cost:' not in install_cost:
        install_cost_decimal = parse_decimal(install_cost)
        if install_cost_decimal and install_cost_decimal == 1.00:
            data['install_cost'] = None
        else:
            data['install_cost'] = install_cost_decimal
    else:
        data['install_cost'] = None
    install_trips = get_cell(143, 9) or get_cell(143, 8) or get_cell(154, 10) or get_cell(154, 9) or get_cell(154, 8)
    install_trips_val = install_trips if install_trips else (get_cell_raw(143, 9) or get_cell_raw(143, 8) or get_cell_raw(154, 10) or get_cell_raw(154, 9) or get_cell_raw(154, 8))
    data['install_trips'] = int(install_trips_val) if install_trips_val and str(install_trips_val).isdigit() else None
    install_days = get_cell(144, 9) or get_cell(144, 8) or get_cell(155, 10) or get_cell(155, 9) or get_cell(155, 8)
    install_days_val = install_days if install_days else (get_cell_raw(144, 9) or get_cell_raw(144, 8) or get_cell_raw(155, 10) or get_cell_raw(155, 9) or get_cell_raw(155, 8))
    data['install_days'] = int(install_days_val) if install_days_val and str(install_days_val).isdigit() else None
    install_net = get_cell(146, 9) or get_cell(146, 8) or get_cell(157, 10) or get_cell(157, 9) or get_cell(157, 8)
    data['install_net_revenue'] = parse_decimal(install_net) if install_net and 'Net Revenue:' not in install_net else None
    
    # TO BE COMPLETED BY ENG (Rows 161-165)
    weld_str = get_cell(161, 3)
    data['eng_weld_in_out'] = parse_bool(weld_str) if weld_str else False
    height_str = get_cell(161, 6)
    data['eng_height_greater_than_20ft'] = parse_bool(height_str) if height_str else False
    crane_str = get_cell(161, 9)
    data['eng_crane_reqd'] = parse_bool(crane_str) if crane_str else False
    
    hi_temp_str = get_cell(163, 3)
    data['eng_hi_temp_htr'] = parse_bool(hi_temp_str) if hi_temp_str else False
    large_hp_str = get_cell(163, 6)
    data['eng_large_hp_or_excessive_ll_pumps'] = parse_bool(large_hp_str) if large_hp_str else False
    generator_str = get_cell(163, 9)
    data['eng_generator_need'] = parse_bool(generator_str) if generator_str else False
    
    testing_str = get_cell(165, 3)
    data['eng_special_testing_reqd'] = parse_bool(testing_str) if testing_str else False
    lift_str = get_cell(165, 6)
    data['eng_extra_forklift_or_scissor_lift_reqd'] = parse_bool(lift_str) if lift_str else False
    
    # Validate critical sections were found
    if not data.get('proposal_number'):
        data['_validation_warnings'].append('Proposal number not found')
    if not data.get('line_items') or len(data['line_items']) == 0:
        data['_validation_warnings'].append('No line items found')
    
    # Return JSON-serializable data
    return data
