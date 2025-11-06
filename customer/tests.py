from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from datetime import datetime, date
from decimal import Decimal
import os
import tempfile
from openpyxl import Workbook

from .models import KOMForm, KOMLineItem, KOMEquipmentRequired
from .utils import parse_kom_excel


def create_test_kom_excel():
    """Create a test KOM Excel file with all fields populated"""
    wb = Workbook()
    ws = wb.active
    
    # Row 2: TO BE COMPLETED BY SALES
    ws.cell(row=2, column=3, value="35371")  # Proposal #
    ws.cell(row=2, column=6, value="9/19/25")  # Proposal Date
    ws.cell(row=2, column=8, value="JOHN O'HEHIR")  # Sales Rep
    ws.cell(row=2, column=11, value="9/29/25")  # Date of OC
    
    # Row 4
    ws.cell(row=4, column=3, value="FOOD")  # Industry
    ws.cell(row=4, column=5, value="POULTRY")  # Industry Subcategory
    ws.cell(row=4, column=8, value="5")  # Discount %
    ws.cell(row=4, column=11, value="4500047876")  # PO#
    
    # Row 6: Commissions
    ws.cell(row=6, column=4, value="100%")  # Comm #1 inside percent
    ws.cell(row=6, column=6, value="JOHN O'HEHIR")  # Comm #1 inside name
    ws.cell(row=6, column=10, value="$34,447.00")  # Comm outside amount
    
    # Row 8
    ws.cell(row=8, column=4, value="0%")  # Comm #2 inside percent
    ws.cell(row=8, column=6, value="")  # Comm #2 inside name
    ws.cell(row=8, column=9, value="Voit Abernathy")  # Comm outside name
    
    # BILL TO - Rows 12-24
    ws.cell(row=12, column=2, value="Bill To Name")
    ws.cell(row=14, column=2, value="704-450-5317")
    ws.cell(row=16, column=2, value="bill@example.com")
    ws.cell(row=18, column=2, value="Bill Company")
    ws.cell(row=20, column=2, value="123 Bill Street")
    ws.cell(row=22, column=3, value="Bill City")
    ws.cell(row=24, column=3, value="WI")
    ws.cell(row=24, column=4, value="53964")
    
    # SHIP TO - Rows 12-24
    ws.cell(row=12, column=6, value="Ship To Name")
    ws.cell(row=14, column=6, value="704-555-1234")
    ws.cell(row=16, column=6, value="ship@example.com")
    ws.cell(row=18, column=6, value="Ship Company")
    ws.cell(row=20, column=6, value="456 Ship Street")
    ws.cell(row=22, column=6, value="Ship City")
    ws.cell(row=24, column=6, value="GA")
    ws.cell(row=24, column=9, value="30643")
    
    # Tax - Row 26
    ws.cell(row=26, column=4, value="NO")  # Tax Exempt
    ws.cell(row=26, column=8, value="YES")  # Exempt Cert in Hand
    ws.cell(row=26, column=11, value="Tax Action Who")
    ws.cell(row=27, column=11, value="10/30/25")  # Tax Action When
    
    # Row 30
    ws.cell(row=30, column=4, value="YES")  # Confirm Tax Status
    ws.cell(row=30, column=6, value="YES")  # Customer in Sage
    
    # Payment Milestones - Rows 33-37
    ws.cell(row=33, column=3, value="Milestone 1 Event")
    ws.cell(row=33, column=4, value=25.5)
    ws.cell(row=33, column=5, value="Net 30")
    ws.cell(row=33, column=6, value="Milestone 1 Notes")
    
    ws.cell(row=34, column=3, value="Milestone 2 Event")
    ws.cell(row=34, column=4, value=50.0)
    ws.cell(row=34, column=5, value="Net 30")
    ws.cell(row=34, column=6, value="Milestone 2 Notes")
    
    # Row 43
    ws.cell(row=43, column=3, value="Consultant Name")
    ws.cell(row=43, column=6, value="Contractor Name")
    
    # Row 45
    ws.cell(row=45, column=3, value="ASAP")
    
    # SHIPPING - Row 48
    ws.cell(row=48, column=3, value="PP&A")
    ws.cell(row=48, column=6, value="NO")
    ws.cell(row=48, column=10, value="Air Freight")
    
    # Row 50
    ws.cell(row=50, column=3, value="Handle with care")
    
    # SPECIFICATIONS - Row 53
    ws.cell(row=53, column=4, value="NO")
    ws.cell(row=53, column=6, value="NO")
    ws.cell(row=53, column=10, value="YES")
    ws.cell(row=54, column=10, value="Rate & Cap Info")
    
    # Row 55
    ws.cell(row=55, column=3, value="NO")
    ws.cell(row=55, column=6, value="BOTH")
    
    # APPROVAL PRINTS - Row 57
    ws.cell(row=57, column=7, value="LL & Mech: 10/17/25 ; Elect: 10/22/25")
    
    # Row 58
    ws.cell(row=58, column=4, value="YES")
    ws.cell(row=58, column=6, value="YES")
    
    # Row 59
    ws.cell(row=59, column=6, value="YES")
    
    # EQUIPMENT - HTR - 1 (Row 61-63)
    ws.cell(row=61, column=4, value=1)
    ws.cell(row=61, column=6, value="TE-100+, NSF")
    ws.cell(row=61, column=9, value="LO NOX - 30PPM")
    ws.cell(row=63, column=2, value="18")
    ws.cell(row=63, column=3, value="x106BTU/hr")
    ws.cell(row=63, column=6, value="PUMP")
    ws.cell(row=63, column=8, value="304")
    
    # HTR - 2 (Row 65-67)
    ws.cell(row=65, column=4, value=2)
    ws.cell(row=65, column=6, value="HTR-2 Type")
    ws.cell(row=65, column=9, value="HTR-2 Emissions")
    ws.cell(row=67, column=2, value="20")
    ws.cell(row=67, column=3, value="x106BTU/hr")
    ws.cell(row=67, column=6, value="GRAV")
    ws.cell(row=67, column=8, value="316")
    
    # STK ECON (Row 69)
    ws.cell(row=69, column=3, value="50 BHP")
    ws.cell(row=69, column=6, value="PUMP")
    ws.cell(row=69, column=8, value="304")
    
    # STACK(S) (Row 71)
    ws.cell(row=71, column=4, value=10)
    ws.cell(row=71, column=6, value="TOTAL")
    ws.cell(row=71, column=9, value="YES")
    
    # HR (Row 73)
    ws.cell(row=73, column=3, value=5)
    ws.cell(row=73, column=5, value=24)
    ws.cell(row=73, column=7, value="100")
    ws.cell(row=73, column=9, value="304")
    
    # TANKS (Rows 77-81)
    ws.cell(row=77, column=3, value="HW")
    ws.cell(row=77, column=5, value=84)
    ws.cell(row=77, column=7, value=15)
    ws.cell(row=77, column=9, value="STD")
    ws.cell(row=77, column=11, value="304")
    
    ws.cell(row=79, column=3, value="CW")
    ws.cell(row=79, column=5, value=96)
    ws.cell(row=79, column=7, value=20)
    ws.cell(row=79, column=9, value="STD")
    ws.cell(row=79, column=11, value="316")
    
    ws.cell(row=81, column=3, value="ST")
    ws.cell(row=81, column=5, value=72)
    ws.cell(row=81, column=7, value=12)
    ws.cell(row=81, column=9, value="STD")
    ws.cell(row=81, column=11, value="304")
    
    # PUMPS (Row 83, 85-91)
    ws.cell(row=83, column=3, value="STANDARD")
    ws.cell(row=83, column=6, value="304")
    
    ws.cell(row=85, column=3, value="HTR DISCH")
    ws.cell(row=85, column=5, value=2)
    ws.cell(row=85, column=7, value=600)
    ws.cell(row=85, column=9, value=80)
    
    ws.cell(row=87, column=3, value="HW")
    ws.cell(row=87, column=5, value=2)
    ws.cell(row=87, column=7, value=400)
    ws.cell(row=87, column=9, value=150)
    
    ws.cell(row=89, column=3, value="RECIRC")
    ws.cell(row=89, column=5, value=1)
    ws.cell(row=89, column=7, value=400)
    ws.cell(row=89, column=9, value=105)
    
    ws.cell(row=91, column=3, value="PUMP4")
    ws.cell(row=91, column=5, value=1)
    ws.cell(row=91, column=7, value=200)
    ws.cell(row=91, column=9, value=50)
    
    # Steam Heaters (Rows 94-96)
    ws.cell(row=94, column=4, value="12x120")
    ws.cell(row=94, column=6, value="304")
    ws.cell(row=94, column=9, value="Ball Valve")
    
    ws.cell(row=96, column=4, value="10x100")
    ws.cell(row=96, column=6, value="316")
    ws.cell(row=96, column=9, value="Gate Valve")
    
    # SOFTENER (Row 99)
    ws.cell(row=99, column=3, value="YES")
    ws.cell(row=99, column=6, value="304")
    ws.cell(row=99, column=9, value="316")
    
    # PANEL(S) (Row 102)
    ws.cell(row=102, column=3, value=1)
    ws.cell(row=102, column=6, value="YES-COMPACT 5069")
    ws.cell(row=102, column=9, value="SPLIT VOLT")
    
    # OTHER (Row 105)
    ws.cell(row=105, column=2, value="Vent Condenser")
    ws.cell(row=105, column=4, value="Shaker Screen")
    
    # UTILITIES (Rows 107-113)
    ws.cell(row=107, column=8, value=2)
    ws.cell(row=109, column=3, value="480/3/60")
    ws.cell(row=109, column=6, value="NG")
    ws.cell(row=109, column=9, value="2-5")
    ws.cell(row=109, column=10, value="(psi)")
    ws.cell(row=111, column=5, value=4)
    ws.cell(row=111, column=9, value="RIGHT")
    ws.cell(row=113, column=8, value="YES")
    
    # OTHER INFO (Rows 116-118)
    ws.cell(row=116, column=3, value="Test notes here")
    ws.cell(row=118, column=2, value="No panel view")
    
    # Project Info (Row 123)
    ws.cell(row=123, column=4, value="Test Project")
    ws.cell(row=123, column=8, value="REPLACEMENT")
    
    # TO BE COMPLETED BY APPS - Line Items (Rows 127-139)
    ws.cell(row=127, column=1, value="35371-01")
    ws.cell(row=127, column=3, value="18 MMBTU DCWH, TE+, Pumped")
    ws.cell(row=127, column=6, value=49550)
    ws.cell(row=127, column=7, value=50000)
    ws.cell(row=127, column=8, value=51000)
    ws.cell(row=127, column=9, value=52000)
    
    ws.cell(row=128, column=1, value="35371-02")
    ws.cell(row=128, column=3, value="Heater FGR Ducting")
    ws.cell(row=128, column=6, value=8447)
    
    ws.cell(row=129, column=1, value="35371-03")
    ws.cell(row=129, column=3, value="Heater Inlet")
    ws.cell(row=129, column=6, value=2243)
    
    # LABOR HOURS (Row 142-145)
    ws.cell(row=142, column=3, value="-")
    ws.cell(row=143, column=3, value=213.0)
    ws.cell(row=144, column=3, value=243.9)
    ws.cell(row=145, column=3, value=101.0)
    
    # EQUIPMENT REQUIRED (Rows 143-145)
    ws.cell(row=143, column=6, value=2)  # Burner qty
    ws.cell(row=143, column=7, value="890-01-012")
    ws.cell(row=143, column=8, value="BURNER, GAS, EB-7")
    
    ws.cell(row=144, column=6, value=1)  # Blower qty
    ws.cell(row=144, column=7, value="892-01-249")
    ws.cell(row=144, column=8, value="BLOWER, 2012S, 30HP, CCWBH")
    
    ws.cell(row=145, column=6, value=153)  # Media qty
    ws.cell(row=145, column=7, value="840-01-002")
    ws.cell(row=145, column=8, value="RING, PACKG CASCADE MNI RNG #3")
    
    # CAPITAL (Rows 149-157)
    ws.cell(row=149, column=3, value="$360,000.00")
    ws.cell(row=151, column=3, value="$164,363.00")
    ws.cell(row=153, column=3, value="$5,000.00")
    ws.cell(row=155, column=4, value="$7,397.00")
    ws.cell(row=156, column=3, value="$2,000.00")
    ws.cell(row=157, column=3, value="$352,603.00")
    
    # INSTALL (Rows 149-157, different columns)
    ws.cell(row=149, column=8, value="$50,000.00")
    ws.cell(row=151, column=8, value="$30,000.00")
    ws.cell(row=154, column=8, value=1)
    ws.cell(row=155, column=8, value=5)
    ws.cell(row=157, column=8, value="$20,000.00")
    
    # TO BE COMPLETED BY ENG (Rows 161-165)
    ws.cell(row=161, column=3, value="YES")
    ws.cell(row=161, column=6, value="YES")
    ws.cell(row=161, column=9, value="YES")
    
    ws.cell(row=163, column=3, value="YES")
    ws.cell(row=163, column=6, value="YES")
    ws.cell(row=163, column=9, value="YES")
    
    ws.cell(row=165, column=3, value="YES")
    ws.cell(row=165, column=6, value="YES")
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name


class KOMImportParserTest(TestCase):
    """Test the KOM Excel parser function directly"""
    
    def setUp(self):
        """Create test Excel file"""
        self.test_file = create_test_kom_excel()
    
    def tearDown(self):
        """Clean up test file"""
        if os.path.exists(self.test_file):
            os.unlink(self.test_file)
    
    def test_parse_sales_info(self):
        """Test that sales information is parsed correctly"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['proposal_number'], '35371')
        self.assertEqual(data['proposal_date'], date(2025, 9, 19))
        self.assertEqual(data['sales_rep'], "JOHN O'HEHIR")
        self.assertEqual(data['date_of_oc'], date(2025, 9, 29))
        self.assertEqual(data['industry'], 'FOOD')
        self.assertEqual(data['industry_subcategory'], 'POULTRY')
        self.assertEqual(data['discount'], Decimal('5'))
        self.assertEqual(data['po_number'], '4500047876')
    
    def test_parse_commissions(self):
        """Test commission fields"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['comm_1_inside_percent'], Decimal('100'))
        self.assertEqual(data['comm_1_inside_name'], "JOHN O'HEHIR")
        self.assertEqual(data['comm_outside_amount'], Decimal('34447.00'))
        self.assertEqual(data['comm_2_inside_percent'], Decimal('0'))
        self.assertEqual(data['comm_outside_name'], 'Voit Abernathy')
    
    def test_parse_bill_to(self):
        """Test Bill To information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['bill_to_name'], 'Bill To Name')
        self.assertEqual(data['bill_to_phone'], '704-450-5317')
        self.assertEqual(data['bill_to_email'], 'bill@example.com')
        self.assertEqual(data['bill_to_company'], 'Bill Company')
        self.assertEqual(data['bill_to_address'], '123 Bill Street')
        self.assertEqual(data['bill_to_city'], 'Bill City')
        self.assertEqual(data['bill_to_state'], 'WI')
        self.assertEqual(data['bill_to_zip'], '53964')
    
    def test_parse_ship_to(self):
        """Test Ship To information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['ship_to_name'], 'Ship To Name')
        self.assertEqual(data['ship_to_phone'], '704-555-1234')
        self.assertEqual(data['ship_to_email'], 'ship@example.com')
        self.assertEqual(data['ship_to_company'], 'Ship Company')
        self.assertEqual(data['ship_to_address'], '456 Ship Street')
        self.assertEqual(data['ship_to_city'], 'Ship City')
        self.assertEqual(data['ship_to_state'], 'GA')
        self.assertEqual(data['ship_to_zip'], '30643')
    
    def test_parse_tax_info(self):
        """Test tax information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertFalse(data['tax_exempt'])  # NO means False
        self.assertTrue(data['exempt_cert_in_hand'])  # YES means True
        self.assertEqual(data['tax_action_who'], 'Tax Action Who')
        self.assertEqual(data['tax_action_when'], date(2025, 10, 30))
        self.assertTrue(data['confirm_tax_status_noted'])
        self.assertTrue(data['customer_in_sage'])
    
    def test_parse_payment_milestones(self):
        """Test payment milestones"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['payment_milestone_1_event'], 'Milestone 1 Event')
        self.assertEqual(data['payment_milestone_1_percent'], Decimal('25.5'))
        self.assertEqual(data['payment_milestone_1_terms'], 'Net 30')
        self.assertEqual(data['payment_milestone_1_notes'], 'Milestone 1 Notes')
        
        self.assertEqual(data['payment_milestone_2_event'], 'Milestone 2 Event')
        self.assertEqual(data['payment_milestone_2_percent'], Decimal('50.0'))
    
    def test_parse_shipping(self):
        """Test shipping information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['freight'], 'PP&A')
        self.assertFalse(data['international'])  # NO means False
        self.assertEqual(data['international_freight_method'], 'Air Freight')
        self.assertEqual(data['shipping_instructions'], 'Handle with care')
    
    def test_parse_specifications(self):
        """Test specifications and terms"""
        data = parse_kom_excel(self.test_file)
        
        self.assertFalse(data['specifications_provided'])
        self.assertFalse(data['specifications_agreed'])
        self.assertTrue(data['liquidated_damages'])
        self.assertEqual(data['liquidated_damages_rate_cap'], 'Rate & Cap Info')
        self.assertFalse(data['passivation'])
        self.assertEqual(data['passivation_in_out_both'], 'BOTH')
    
    def test_parse_approval_prints(self):
        """Test approval prints"""
        data = parse_kom_excel(self.test_file)
        
        self.assertTrue(data['approval_prints_required'])
        self.assertTrue(data['approval_prints_electrical'])
        self.assertIn('LL & Mech: 10/17/25', data['approval_prints_ll_mech'])
        self.assertIn('10/22/25', data['approval_prints_elect'])
        self.assertTrue(data['engineering_order_prior_to_approval'])
    
    def test_parse_equipment_heaters(self):
        """Test heater equipment"""
        data = parse_kom_excel(self.test_file)
        
        # HTR - 1
        self.assertEqual(data['htr_1_qty'], Decimal('1'))
        self.assertEqual(data['htr_1_type'], "TE-100+, NSF")
        self.assertEqual(data['htr_1_emissions'], 'LO NOX - 30PPM')
        self.assertIn('18', data['htr_1_size'])
        self.assertEqual(data['htr_1_pump_grav'], 'PUMP')
        self.assertEqual(data['htr_1_material'], '304')
        
        # HTR - 2
        self.assertEqual(data['htr_2_qty'], Decimal('2'))
        self.assertEqual(data['htr_2_type'], 'HTR-2 Type')
        self.assertEqual(data['htr_2_emissions'], 'HTR-2 Emissions')
        self.assertIn('20', data['htr_2_size'])
        self.assertEqual(data['htr_2_pump_grav'], 'GRAV')
        self.assertEqual(data['htr_2_material'], '316')
    
    def test_parse_stack_economizer(self):
        """Test stack economizer"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['stk_econ_size_bhp'], '50 BHP')
        self.assertEqual(data['stk_econ_pump_grav'], 'PUMP')
        self.assertEqual(data['stk_econ_material'], '304')
    
    def test_parse_stack(self):
        """Test stack information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['stack_length_ft'], Decimal('10'))
        self.assertEqual(data['stack_total'], 'TOTAL')
        self.assertTrue(data['stack_caps'])
    
    def test_parse_hr(self):
        """Test HR information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['hr_sections'], Decimal('5'))
        self.assertEqual(data['hr_diam_in'], Decimal('24'))
        self.assertEqual(data['hr_tubes'], '100')
        self.assertEqual(data['hr_material'], '304')
    
    def test_parse_tanks(self):
        """Test tank information"""
        data = parse_kom_excel(self.test_file)
        
        # Tank 1
        self.assertEqual(data['tank_1_type'], 'HW')
        self.assertEqual(data['tank_1_dia_in'], Decimal('84'))
        self.assertEqual(data['tank_1_ht_ft'], Decimal('15'))
        self.assertEqual(data['tank_1_ga'], 'STD')
        self.assertEqual(data['tank_1_material'], '304')
        
        # Tank 2
        self.assertEqual(data['tank_2_type'], 'CW')
        self.assertEqual(data['tank_2_dia_in'], Decimal('96'))
        self.assertEqual(data['tank_2_ht_ft'], Decimal('20'))
        self.assertEqual(data['tank_2_ga'], 'STD')
        self.assertEqual(data['tank_2_material'], '316')
        
        # Tank 3
        self.assertEqual(data['tank_3_type'], 'ST')
        self.assertEqual(data['tank_3_dia_in'], Decimal('72'))
        self.assertEqual(data['tank_3_ht_ft'], Decimal('12'))
        self.assertEqual(data['tank_3_ga'], 'STD')
        self.assertEqual(data['tank_3_material'], '304')
    
    def test_parse_pumps(self):
        """Test pump information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['pump_packaging'], 'STANDARD')
        self.assertEqual(data['pump_piping_material'], '304')
        
        # Pump 1
        self.assertEqual(data['pump_1_type'], 'HTR DISCH')
        self.assertEqual(data['pump_1_qty'], Decimal('2'))
        self.assertEqual(data['pump_1_flow_gpm'], Decimal('600'))
        self.assertEqual(data['pump_1_tdh_ft'], Decimal('80'))
        
        # Pump 2
        self.assertEqual(data['pump_2_type'], 'HW')
        self.assertEqual(data['pump_2_qty'], Decimal('2'))
        self.assertEqual(data['pump_2_flow_gpm'], Decimal('400'))
        self.assertEqual(data['pump_2_tdh_ft'], Decimal('150'))
        
        # Pump 3
        self.assertEqual(data['pump_3_type'], 'RECIRC')
        self.assertEqual(data['pump_3_qty'], Decimal('1'))
        self.assertEqual(data['pump_3_flow_gpm'], Decimal('400'))
        self.assertEqual(data['pump_3_tdh_ft'], Decimal('105'))
        
        # Pump 4
        self.assertEqual(data['pump_4_type'], 'PUMP4')
        self.assertEqual(data['pump_4_qty'], Decimal('1'))
        self.assertEqual(data['pump_4_flow_gpm'], Decimal('200'))
        self.assertEqual(data['pump_4_tdh_ft'], Decimal('50'))
    
    def test_parse_steam_heaters(self):
        """Test steam heater information"""
        data = parse_kom_excel(self.test_file)
        
        # Steam Heater 1
        self.assertEqual(data['steam_heater_1_dia_in'], Decimal('12'))
        self.assertEqual(data['steam_heater_1_length_in'], Decimal('120'))
        self.assertEqual(data['steam_heater_1_material'], '304')
        self.assertEqual(data['steam_heater_1_valve_type'], 'Ball Valve')
        
        # Steam Heater 2
        self.assertEqual(data['steam_heater_2_dia_in'], Decimal('10'))
        self.assertEqual(data['steam_heater_2_length_in'], Decimal('100'))
        self.assertEqual(data['steam_heater_2_material'], '316')
        self.assertEqual(data['steam_heater_2_valve_type'], 'Gate Valve')
    
    def test_parse_softener(self):
        """Test softener information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertTrue(data['softener_asme_coded'])
        self.assertEqual(data['softener_tank_material'], '304')
        self.assertEqual(data['softener_face_plumbing_material'], '316')
    
    def test_parse_panel(self):
        """Test panel information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['panel_qty'], Decimal('1'))
        self.assertEqual(data['panel_plc'], 'YES-COMPACT 5069')
        self.assertEqual(data['panel_split_volt'], 'SPLIT VOLT')
    
    def test_parse_other_equipment(self):
        """Test other equipment"""
        data = parse_kom_excel(self.test_file)
        
        self.assertTrue(data['other_vent_condenser'])
        self.assertTrue(data['other_shaker_screen'])
    
    def test_parse_utilities(self):
        """Test utilities information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['city_water_meter_in'], Decimal('2'))
        self.assertEqual(data['electrical'], '480/3/60')
        self.assertEqual(data['fuel_type'], 'NG')
        self.assertIn('2-5', data['gas_pressure_psi'])
        self.assertEqual(data['onsite_gas_supply_diameter_in'], Decimal('4'))
        self.assertEqual(data['gas_train_orientation'], 'RIGHT')
        self.assertTrue(data['utilities_match_proposal'])
    
    def test_parse_other_info(self):
        """Test other info"""
        data = parse_kom_excel(self.test_file)
        
        self.assertIn('Test notes here', data['notes'])
        self.assertIn('No panel view', data['notes'])
        self.assertEqual(data['project_name'], 'Test Project')
        self.assertEqual(data['project_type'], 'REPLACEMENT')
    
    def test_parse_line_items(self):
        """Test line items"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(len(data['line_items']), 3)
        
        item1 = data['line_items'][0]
        self.assertEqual(item1['item_number'], '35371-01')
        self.assertEqual(item1['description'], '18 MMBTU DCWH, TE+, Pumped')
        self.assertEqual(item1['value_1'], Decimal('49550'))
        self.assertEqual(item1['value_2'], Decimal('50000'))
        self.assertEqual(item1['value_3'], Decimal('51000'))
        self.assertEqual(item1['value_4'], Decimal('52000'))
        
        item2 = data['line_items'][1]
        self.assertEqual(item2['item_number'], '35371-02')
        self.assertEqual(item2['description'], 'Heater FGR Ducting')
        self.assertEqual(item2['value_1'], Decimal('8447'))
    
    def test_parse_labor_hours(self):
        """Test labor hours"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['labor_hr'], '-')
        self.assertEqual(data['labor_pkg'], Decimal('213.0'))
        self.assertEqual(data['labor_fab'], Decimal('243.9'))
        self.assertEqual(data['labor_wiring'], Decimal('101.0'))
    
    def test_parse_equipment_required(self):
        """Test equipment required"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(len(data['equipment_required']), 3)
        
        burner = data['equipment_required'][0]
        self.assertEqual(burner['equipment_type'], 'Burner')
        self.assertEqual(burner['qty'], 2)
        self.assertEqual(burner['kn_number'], '890-01-012')
        self.assertEqual(burner['description'], 'BURNER, GAS, EB-7')
        
        blower = data['equipment_required'][1]
        self.assertEqual(blower['equipment_type'], 'Blower')
        self.assertEqual(blower['qty'], 1)
        self.assertEqual(blower['kn_number'], '892-01-249')
        
        media = data['equipment_required'][2]
        self.assertEqual(media['equipment_type'], 'Media')
        self.assertEqual(media['qty'], 153)
        self.assertEqual(media['kn_number'], '840-01-002')
    
    def test_parse_capital(self):
        """Test capital information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['capital_sell_price'], Decimal('360000.00'))
        self.assertEqual(data['capital_equip_cost'], Decimal('164363.00'))
        self.assertEqual(data['capital_freight'], Decimal('5000.00'))
        self.assertEqual(data['capital_startup_cost'], Decimal('7397.00'))
        self.assertEqual(data['capital_protect_cost'], Decimal('2000.00'))
        self.assertEqual(data['capital_net_revenue'], Decimal('352603.00'))
    
    def test_parse_install(self):
        """Test install information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertEqual(data['install_sell_price'], Decimal('50000.00'))
        self.assertEqual(data['install_cost'], Decimal('30000.00'))
        self.assertEqual(data['install_trips'], 1)
        self.assertEqual(data['install_days'], 5)
        self.assertEqual(data['install_net_revenue'], Decimal('20000.00'))
    
    def test_parse_engineering(self):
        """Test engineering information"""
        data = parse_kom_excel(self.test_file)
        
        self.assertTrue(data['eng_weld_in_out'])
        self.assertTrue(data['eng_height_greater_than_20ft'])
        self.assertTrue(data['eng_crane_reqd'])
        self.assertTrue(data['eng_hi_temp_htr'])
        self.assertTrue(data['eng_large_hp_or_excessive_ll_pumps'])
        self.assertTrue(data['eng_generator_need'])
        self.assertTrue(data['eng_special_testing_reqd'])
        self.assertTrue(data['eng_extra_forklift_or_scissor_lift_reqd'])


class KOMImportViewTest(TestCase):
    """Test the KOM import view"""
    
    def setUp(self):
        """Create test user and Excel file"""
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123',
            is_superuser=True
        )
        self.client = Client()
        self.client.login(username='testuser', password='testpass123')
        self.test_file = create_test_kom_excel()
    
    def tearDown(self):
        """Clean up test file"""
        if os.path.exists(self.test_file):
            os.unlink(self.test_file)
    
    def test_import_kom_form(self):
        """Test that importing a KOM file creates all models correctly"""
        with open(self.test_file, 'rb') as f:
            response = self.client.post('/customer/kom/import/', {
                'kom_file': SimpleUploadedFile(
                    'test_kom.xlsx',
                    f.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            })
        
        # Should redirect to detail page
        self.assertEqual(response.status_code, 302)
        
        # Check that KOMForm was created
        kom_form = KOMForm.objects.get(proposal_number='35371')
        self.assertEqual(kom_form.sales_rep, "JOHN O'HEHIR")
        self.assertEqual(kom_form.project_name, 'Test Project')
        self.assertEqual(kom_form.created_by, self.user)
        
        # Check line items were created
        line_items = kom_form.line_items.all()
        self.assertEqual(line_items.count(), 3)
        self.assertTrue(line_items.filter(item_number='35371-01').exists())
        
        # Check equipment required was created
        equipment = kom_form.equipment_required.all()
        self.assertEqual(equipment.count(), 3)
        self.assertTrue(equipment.filter(equipment_type='Burner').exists())
        self.assertTrue(equipment.filter(equipment_type='Blower').exists())
        self.assertTrue(equipment.filter(equipment_type='Media').exists())
    
    def test_import_kom_all_fields_present(self):
        """Test that ALL fields from the Excel are imported"""
        with open(self.test_file, 'rb') as f:
            self.client.post('/customer/kom/import/', {
                'kom_file': SimpleUploadedFile(
                    'test_kom.xlsx',
                    f.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            })
        
        kom_form = KOMForm.objects.get(proposal_number='35371')
        
        # Verify all major sections are populated
        self.assertTrue(kom_form.proposal_number)
        self.assertTrue(kom_form.sales_rep)
        self.assertTrue(kom_form.bill_to_company)
        self.assertTrue(kom_form.ship_to_company)
        self.assertTrue(kom_form.project_name)
        self.assertTrue(kom_form.line_items.exists())
        self.assertTrue(kom_form.equipment_required.exists())
        
        # Verify specific fields that are often missed
        self.assertIsNotNone(kom_form.tax_action_when)
        self.assertTrue(kom_form.payment_milestone_1_event)
        self.assertTrue(kom_form.consultant or kom_form.contractor_name)
        self.assertTrue(kom_form.htr_1_type)
        self.assertTrue(kom_form.tank_1_type)
        self.assertTrue(kom_form.pump_1_type)
        self.assertTrue(kom_form.labor_pkg is not None)
