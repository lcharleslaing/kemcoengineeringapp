"""
GUI tool to analyze and compare KOM Excel file structures
Uses tkinter for file selection and results display
"""
import os
import sys
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from collections import defaultdict
from openpyxl import load_workbook
import json
from datetime import datetime


def get_cell_value(ws, row, col):
    """Safely get cell value"""
    try:
        val = ws.cell(row=row, column=col).value
        if val is None:
            return ''
        return str(val).strip()
    except:
        return ''


def find_section_headers(ws, max_row=200):
    """Find all section headers in the worksheet"""
    sections = []
    for row in range(1, max_row):
        for col in range(1, 15):
            val = get_cell_value(ws, row, col)
            if val:
                val_upper = val.upper()
                section_keywords = [
                    'TO BE COMPLETED BY SALES',
                    'TO BE COMPLETED BY APPS',
                    'TO BE COMPLETED BY ENG',
                    'TO BE COMPLETED BY ACCT',
                    'BILL TO',
                    'SHIP TO',
                    'TAX',
                    'PAYMENT MILESTONES',
                    'SHIPPING',
                    'SPECIFICATIONS',
                    'APPROVAL PRINTS',
                    'EQUIPMENT',
                    'HTR - 1',
                    'HTR - 2',
                    'STK ECON',
                    'STACK',
                    'HR',
                    'TANKS',
                    'PUMPS',
                    'STEAM HEATERS',
                    'SOFTENER',
                    'PANEL',
                    'OTHER',
                    'UTILITIES',
                    'LABOR HOURS',
                    'CAPITAL',
                    'INSTALL',
                    'PROJECT NAME',
                ]
                for keyword in section_keywords:
                    if keyword in val_upper:
                        sections.append({
                            'row': row,
                            'col': col,
                            'text': val,
                            'keyword': keyword
                        })
                        break
    return sections


def analyze_tanks_section(ws, tanks_row):
    """Analyze the tanks section structure"""
    tanks_data = []
    for row in range(tanks_row, tanks_row + 10):
        marker = get_cell_value(ws, row, 1)
        if marker and marker.startswith('#'):
            tank_info = {
                'row': row,
                'marker': marker,
                'cells': {}
            }
            for col in range(1, 16):
                val = get_cell_value(ws, row, col)
                if val and val not in ['', '#1:', '#2:', '#3:', 'Type:', 'Dia (in):', 'Ht (ft):', 'GA:', "Mat'l:"]:
                    tank_info['cells'][col] = val
            tanks_data.append(tank_info)
    return tanks_data


def analyze_line_items_section(ws, line_items_start_row):
    """Analyze line items structure"""
    line_items = []
    for row in range(line_items_start_row, line_items_start_row + 20):
        item_num = get_cell_value(ws, row, 1)
        if item_num and item_num not in ['ITEM', 'DESCRIPTION', 'TO BE COMPLETED BY APPS', 'LABOR HOURS']:
            if item_num and not item_num.endswith(':') and len(item_num) > 0:
                item_data = {
                    'row': row,
                    'item_number': item_num,
                    'description_col': None,
                    'value_cols': []
                }
                desc = get_cell_value(ws, row, 4)
                if desc:
                    item_data['description_col'] = 4
                    item_data['description'] = desc
                
                for col in range(6, 13):
                    val = ws.cell(row=row, column=col).value
                    if val is not None:
                        try:
                            float(str(val).replace('$', '').replace(',', ''))
                            item_data['value_cols'].append({
                                'col': col,
                                'value': val
                            })
                        except:
                            pass
                
                if item_data['value_cols']:
                    line_items.append(item_data)
    return line_items


def analyze_file(file_path, progress_callback=None):
    """Analyze a single KOM Excel file"""
    if progress_callback:
        progress_callback(f"Analyzing: {os.path.basename(file_path)}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        analysis = {
            'filename': os.path.basename(file_path),
            'filepath': file_path,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'sections': find_section_headers(ws),
            'tanks': [],
            'line_items': [],
            'sample_data': {}
        }
        
        for section in analysis['sections']:
            if 'TANKS' in section['keyword']:
                tanks_data = analyze_tanks_section(ws, section['row'])
                analysis['tanks'] = tanks_data
        
        for section in analysis['sections']:
            if 'TO BE COMPLETED BY APPS' in section['keyword']:
                line_items = analyze_line_items_section(ws, section['row'] + 2)
                analysis['line_items'] = line_items
        
        analysis['sample_data'] = {
            'proposal_number': get_cell_value(ws, 2, 3),
            'htr1_type': get_cell_value(ws, 61, 7),
            'htr1_size': get_cell_value(ws, 63, 3),
            'labor_pkg': get_cell_value(ws, 143, 4),
            'equipment_qty': get_cell_value(ws, 143, 7),
        }
        
        return analysis
        
    except Exception as e:
        if progress_callback:
            progress_callback(f"ERROR: {str(e)}")
        return None


def compare_files(analyses):
    """Compare multiple file analyses"""
    if not analyses:
        return {}
    
    comparison = {
        'section_positions': defaultdict(list),
        'tank_columns': defaultdict(set),
        'value_columns': defaultdict(set),
        'desc_columns': set(),
    }
    
    for analysis in analyses:
        for section in analysis['sections']:
            comparison['section_positions'][section['keyword']].append({
                'file': analysis['filename'],
                'row': section['row'],
                'col': section['col']
            })
        
        for tank in analysis['tanks']:
            for col in tank['cells'].keys():
                comparison['tank_columns'][tank['marker']].add(col)
        
        for item in analysis['line_items']:
            if item['description_col']:
                comparison['desc_columns'].add(item['description_col'])
            for val in item['value_cols']:
                comparison['value_columns'][val['col']].add(analysis['filename'])
    
    return comparison


class KOMAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("KOM File Structure Analyzer")
        self.root.geometry("1200x800")
        
        self.analyses = []
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.file_listbox = tk.Listbox(file_frame, height=4, width=80)
        self.file_listbox.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        btn_frame = ttk.Frame(file_frame)
        btn_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        ttk.Button(btn_frame, text="Add Files...", command=self.add_files).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(btn_frame, text="Remove Selected", command=self.remove_file).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(btn_frame, text="Clear All", command=self.clear_files).grid(row=0, column=2, padx=(0, 5))
        ttk.Button(btn_frame, text="Analyze Files", command=self.analyze_files).grid(row=0, column=3, padx=(0, 5))
        ttk.Button(btn_frame, text="Save Report", command=self.save_report).grid(row=0, column=4)
        
        # Results frame with notebook (tabs)
        results_frame = ttk.LabelFrame(main_frame, text="Analysis Results", padding="10")
        results_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        main_frame.rowconfigure(1, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_frame.rowconfigure(0, weight=1)
        results_frame.columnconfigure(0, weight=1)
        
        # Progress/Status frame
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        self.status_label = ttk.Label(status_frame, text="Ready. Select files and click 'Analyze Files'.")
        self.status_label.grid(row=0, column=0, sticky=tk.W)
        
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate')
        self.progress.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
        status_frame.columnconfigure(1, weight=1)
    
    def add_files(self):
        """Open file picker to select KOM Excel files"""
        files = filedialog.askopenfilenames(
            title="Select KOM Excel Files",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        for file_path in files:
            if file_path not in self.file_listbox.get(0, tk.END):
                self.file_listbox.insert(tk.END, file_path)
        
        self.update_status(f"Added {len(files)} file(s). Total: {self.file_listbox.size()}")
    
    def remove_file(self):
        """Remove selected file from list"""
        selection = self.file_listbox.curselection()
        for index in reversed(selection):
            self.file_listbox.delete(index)
        self.update_status("File removed from list")
    
    def clear_files(self):
        """Clear all files from list"""
        self.file_listbox.delete(0, tk.END)
        self.analyses = []
        self.update_status("File list cleared")
    
    def update_status(self, message):
        """Update status label"""
        self.status_label.config(text=message)
        self.root.update_idletasks()
    
    def analyze_files(self):
        """Analyze all selected files"""
        files = list(self.file_listbox.get(0, tk.END))
        
        if not files:
            messagebox.showwarning("No Files", "Please select at least one file to analyze.")
            return
        
        self.analyses = []
        self.progress.start()
        
        try:
            # Clear existing tabs
            for tab in self.notebook.tabs():
                self.notebook.forget(tab)
            
            # Analyze each file
            for i, file_path in enumerate(files):
                self.update_status(f"Analyzing {i+1}/{len(files)}: {os.path.basename(file_path)}")
                analysis = analyze_file(file_path, self.update_status)
                if analysis:
                    self.analyses.append(analysis)
                    self.create_file_tab(analysis)
            
            # Create comparison tab if multiple files
            if len(self.analyses) > 1:
                self.create_comparison_tab()
            
            self.update_status(f"Analysis complete! Analyzed {len(self.analyses)} file(s).")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error during analysis: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.progress.stop()
    
    def create_file_tab(self, analysis):
        """Create a tab showing analysis for a single file"""
        frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(frame, text=analysis['filename'])
        
        text = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=100, height=40)
        text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        
        output = []
        output.append("=" * 80)
        output.append(f"FILE: {analysis['filename']}")
        output.append("=" * 80)
        output.append(f"\nFile Path: {analysis['filepath']}")
        output.append(f"Dimensions: {analysis['max_row']} rows x {analysis['max_col']} columns")
        
        output.append("\n" + "=" * 80)
        output.append("SECTION HEADERS")
        output.append("=" * 80)
        for section in sorted(analysis['sections'], key=lambda x: x['row']):
            output.append(f"  Row {section['row']:3d}, Col {section['col']:2d}: {section['keyword']:40s} ({section['text']})")
        
        if analysis['tanks']:
            output.append("\n" + "=" * 80)
            output.append("TANKS STRUCTURE")
            output.append("=" * 80)
            for tank in analysis['tanks']:
                output.append(f"\n  {tank['marker']} at Row {tank['row']}:")
                for col, val in sorted(tank['cells'].items()):
                    output.append(f"    Column {col:2d}: {val}")
        
        if analysis['line_items']:
            output.append("\n" + "=" * 80)
            output.append("LINE ITEMS STRUCTURE")
            output.append("=" * 80)
            output.append(f"  Found {len(analysis['line_items'])} line items")
            output.append(f"  Description column: {analysis['line_items'][0].get('description_col', 'N/A')}")
            value_cols = set()
            for item in analysis['line_items']:
                for val in item['value_cols']:
                    value_cols.add(val['col'])
            output.append(f"  Value columns: {sorted(value_cols)}")
            output.append("\n  Sample items:")
            for item in analysis['line_items'][:5]:
                output.append(f"    Row {item['row']}: {item['item_number']}")
                output.append(f"      Description: {item.get('description', 'N/A')[:50]}")
                value_strs = [f"Col {v['col']}={v['value']}" for v in item['value_cols']]
                output.append(f"      Values: {value_strs}")
        
        output.append("\n" + "=" * 80)
        output.append("SAMPLE DATA POINTS")
        output.append("=" * 80)
        for key, value in analysis['sample_data'].items():
            output.append(f"  {key:20s}: {value}")
        
        text.insert('1.0', '\n'.join(output))
        text.config(state=tk.DISABLED)
    
    def create_comparison_tab(self):
        """Create a tab showing comparison of all files"""
        frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(frame, text="Comparison", state='normal')
        
        text = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=100, height=40)
        text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        
        comparison = compare_files(self.analyses)
        
        output = []
        output.append("=" * 80)
        output.append("FILE COMPARISON SUMMARY")
        output.append("=" * 80)
        output.append(f"\nComparing {len(self.analyses)} files:")
        for analysis in self.analyses:
            output.append(f"  - {analysis['filename']}")
        
        output.append("\n" + "=" * 80)
        output.append("SECTION POSITIONS")
        output.append("=" * 80)
        for keyword, positions in sorted(comparison['section_positions'].items()):
            rows = [p['row'] for p in positions]
            cols = [p['col'] for p in positions]
            if len(set(rows)) == 1 and len(set(cols)) == 1:
                output.append(f"\n{keyword:40s} ✓ CONSISTENT")
                output.append(f"  Row {rows[0]}, Col {cols[0]}")
            else:
                output.append(f"\n{keyword:40s} ⚠ VARIES")
                for pos in positions:
                    output.append(f"  {pos['file']:40s} Row {pos['row']:3d}, Col {pos['col']:2d}")
        
        output.append("\n" + "=" * 80)
        output.append("TANKS STRUCTURE")
        output.append("=" * 80)
        for marker, cols in sorted(comparison['tank_columns'].items()):
            output.append(f"\n{marker:10s} Data columns: {sorted(cols)}")
        
        output.append("\n" + "=" * 80)
        output.append("LINE ITEMS STRUCTURE")
        output.append("=" * 80)
        output.append(f"Description columns: {sorted(comparison['desc_columns'])}")
        output.append("\nValue columns:")
        for col in sorted(comparison['value_columns'].keys()):
            files = comparison['value_columns'][col]
            if len(files) == len(self.analyses):
                output.append(f"  Col {col:2d}: ✓ CONSISTENT (all {len(files)} files)")
            else:
                output.append(f"  Col {col:2d}: ⚠ Used in {len(files)}/{len(self.analyses)} files")
                for filename in files:
                    output.append(f"      - {filename}")
        
        text.insert('1.0', '\n'.join(output))
        text.config(state=tk.DISABLED)
    
    def save_report(self):
        """Save analysis report to JSON file"""
        if not self.analyses:
            messagebox.showwarning("No Data", "Please analyze files first.")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Save Analysis Report",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                report = {
                    'timestamp': datetime.now().isoformat(),
                    'files_analyzed': len(self.analyses),
                    'analyses': self.analyses,
                    'comparison': compare_files(self.analyses) if len(self.analyses) > 1 else None
                }
                
                with open(filename, 'w') as f:
                    json.dump(report, f, indent=2, default=str)
                
                messagebox.showinfo("Success", f"Report saved to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error saving report: {str(e)}")


def main():
    """Main function to run the GUI"""
    root = tk.Tk()
    app = KOMAnalyzerGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()

